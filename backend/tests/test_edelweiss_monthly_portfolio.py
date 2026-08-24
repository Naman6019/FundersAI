from __future__ import annotations

from datetime import date

from openpyxl import Workbook

from app.mf_ingestion.downloaders.amc_downloader import (
    _edelweiss_monthly_portfolio_documents_from_candidates,
)
from app.mf_ingestion.parsers.adapters.edelweiss_adapter import EdelweissAdapter
from app.mf_ingestion.parsers.base_parser import ParseContext
from app.mf_ingestion.parsers.holdings_parser import HoldingsParser
from app.mf_ingestion.services.parsing_service import _snapshot_matches_amc
from app.mf_ingestion.sources.registry import get_source


EDELWEISS_PORTFOLIO_URL = "https://www.edelweissmf.com/statutory/portfolio-of-schemes"
JULY_WORKBOOK_URL = (
    "https://www.edelweissmf.com/Files/MF/Statutory/Portfolio_of_schemes/"
    "Monthly_Portfolio_and_RiskoMeter/EDEL_Portfolio_Monthly_Notes_31Jul2026.xlsx"
)


def test_edelweiss_monthly_browser_candidates_keep_only_full_monthly_workbook() -> None:
    documents = _edelweiss_monthly_portfolio_documents_from_candidates(
        get_source("edelweiss"),
        listing_url=EDELWEISS_PORTFOLIO_URL,
        candidates=[
            ("Weekly Portfolio of selected schemes - August 14, 2026", "https://www.edelweissmf.com/weekly.xlsx"),
            ("Fortnightly Portfolio of Debt Schemes - July 2026", "https://www.edelweissmf.com/fortnightly.xlsx"),
            ("Monthly Portfolio - July 31, 2026", JULY_WORKBOOK_URL),
        ],
    )

    assert len(documents) == 1
    assert documents[0].document_type == "portfolio_disclosure"
    assert documents[0].report_month == date(2026, 7, 1)
    assert documents[0].file_ext == ".xlsx"
    assert documents[0].url == JULY_WORKBOOK_URL


def test_edelweiss_workbook_title_and_fractional_allocations_are_normalized(tmp_path) -> None:
    workbook = Workbook()
    index = workbook.active
    index.title = "Index"
    index.append(["Portfolio statement"])
    index.append(["Fund Id", "Fund Desc"])
    index.append(["EEDGEF", "Edelweiss Large Cap Fund"])

    sheet = workbook.create_sheet("EEDGEF")
    sheet.title = "EEDGEF"
    sheet.append(["PORTFOLIO STATEMENT OF EDELWEISS LARGE CAP FUND AS ON JULY 31, 2026"])
    sheet.append(["(An open ended equity scheme)"])
    sheet.append([])
    sheet.append(
        [
            "Name of the Instrument",
            "ISIN",
            "Rating/Industry",
            "Quantity",
            "Market/Fair Value(Rs. In Lacs)",
            "% to Net Assets",
        ]
    )
    sheet.append(["HDFC Bank Ltd.", "INE040A01034", "Banks", 1, 60, 0.60])
    sheet.append(["ICICI Bank Ltd.", "INE090A01021", "Banks", 1, 40, 0.40])
    sheet.append(["Grand Total", None, None, None, 100, 1.00])

    index.append(["EENLMG", "Edelweiss Nifty LargeMidcap250 Plus 8-13 yr G-Sec 70:30 Index Fund"])
    abbreviated_sheet = workbook.create_sheet("EENLMG")
    abbreviated_sheet.append(["PORTFOLIO STATEMENT OF EDEL NY LMCAP250 PL 8 13 YR GS 70 30 IDX AS ON JULY 31, 2026"])
    abbreviated_sheet.append(["Name of the Instrument", "ISIN", "Rating/Industry", "% to Net Assets"])
    abbreviated_sheet.append(["Government Security", "IN0020220148", "Sovereign", 1.00])
    abbreviated_sheet.append(["Grand Total", None, None, 1.00])
    file_path = tmp_path / "edelweiss-july-2026.xlsx"
    workbook.save(file_path)

    result = HoldingsParser(EdelweissAdapter()).parse_batch(
        str(file_path),
        ParseContext(
            source_document_id="edelweiss-july",
            source_url=JULY_WORKBOOK_URL,
            report_month=date(2026, 7, 1),
        ),
    )

    assert result.successful_sources == 2
    assert result.empty_sources == 1
    assert len(result.records) == 2
    records = {record.scheme_name: record for record in result.records}
    large_cap = records["Edelweiss Large Cap Fund"]
    assert large_cap.metrics["total_percent_aum"] == 100.0
    assert [item["percent_aum"] for item in large_cap.holdings] == [60.0, 40.0]
    assert "Edelweiss Nifty LargeMidcap250 Plus 8-13 yr G-Sec 70:30 Index Fund" in records


def test_edelweiss_mapping_alias_accepts_its_official_snapshot_name() -> None:
    source = get_source("edelweiss")

    assert source.factsheet_contains_holdings is False
    assert source.portfolio_disclosure_page_url == EDELWEISS_PORTFOLIO_URL
    assert _snapshot_matches_amc("edelweiss", "Edelweiss Mutual Fund") is True
