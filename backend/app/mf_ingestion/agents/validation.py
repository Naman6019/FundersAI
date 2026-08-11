from __future__ import annotations

from datetime import date, timedelta
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from urllib.parse import urlsplit
import re
import zipfile

from app.mf_ingestion.downloaders.base_downloader import (
    DiscoveredDocument,
    DownloadedDocument,
    local_file_sources_allowed,
)
from app.mf_ingestion.sources.registry import AMCDocumentSource

SUPPORTED_DOCUMENT_TYPES = {"factsheet", "portfolio_disclosure"}
SUPPORTED_EXTENSIONS = {".pdf", ".xls", ".xlsx", ".xlsm", ".csv", ".zip", ".html", ".htm"}
UTI_PORTFOLIO_MONTH_PATTERN = re.compile(
    r"portfolio\s+disclosure\s+as\s+of\s+(?P<day>\d{1,2})[/-](?P<month>\d{1,2})[/-](?P<year>20\d{2})",
    re.IGNORECASE,
)

EXTRA_OFFICIAL_HOST_SUFFIXES: dict[str, tuple[str, ...]] = {
    "aditya_birla": ("adityabirlacapital.com",),
    "axis": ("axismf.com",),
    "dsp": ("dspim.com",),
    "hdfc": ("hdfcfund.com",),
    "icici": ("icicipruamc.com",),
    "kotak": ("kotakmf.com",),
    "mirae": ("miraeassetmf.co.in",),
    "nippon": ("nipponindiaim.com",),
    "ppfas": ("ppfas.com",),
    "sbi": ("sbimf.com",),
    "uti": ("utimf.com", "d3ce1o48hc5oli.cloudfront.net"),
    "tata": ("tatamutualfund.com",),
    "bandhan": ("bandhanmutual.com",),
    "edelweiss": ("edelweissmf.com",),
    "invesco": ("invescomutualfund.com",),
    "hsbc": ("assetmanagement.hsbc.co.in",),
}


def validate_candidate(
    source: AMCDocumentSource,
    document: DiscoveredDocument,
    *,
    expected_month: date | None,
    expected_month_grace_days: int = 14,
    observed_on: date | None = None,
) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    document_type = str(document.document_type or "").strip().lower()

    if document_type not in SUPPORTED_DOCUMENT_TYPES:
        errors.append(f"unsupported_document_type:{document_type or 'missing'}")

    parsed = urlsplit(str(document.url or ""))
    if parsed.scheme == "file":
        if not local_file_sources_allowed():
            errors.append("local_file_source_not_allowed")
    elif parsed.scheme not in {"http", "https"} or not parsed.hostname:
        errors.append("invalid_source_url")
    elif not _is_official_host(source, parsed.hostname):
        errors.append(f"non_official_host:{parsed.hostname.lower()}")

    extension = _normalize_extension(document.file_ext or document.url)
    if extension not in SUPPORTED_EXTENSIONS:
        errors.append(f"unsupported_file_type:{extension or 'missing'}")

    if document.report_month is None:
        warnings.append("report_month_unknown")
    elif expected_month and _month_index(document.report_month) != _month_index(expected_month):
        next_month = (
            date(expected_month.year + 1, 1, 1)
            if expected_month.month == 12
            else date(expected_month.year, expected_month.month + 1, 1)
        )
        grace_deadline = next_month + timedelta(days=max(int(expected_month_grace_days), 0) - 1)
        if (observed_on or date.today()) <= grace_deadline:
            warnings.append(f"report_month_pending_expected:{document.report_month.isoformat()}")
        elif _month_index(document.report_month) < _month_index(expected_month):
            warnings.append(f"report_month_before_expected:{document.report_month.isoformat()}")
        else:
            warnings.append(f"report_month_after_expected:{document.report_month.isoformat()}")

    if not str(document.discovery_page_url or "").strip():
        warnings.append("discovery_page_unknown")

    return errors, warnings


def validate_download(
    source: AMCDocumentSource,
    downloaded: DownloadedDocument,
) -> list[str]:
    errors: list[str] = []
    parsed = urlsplit(str(downloaded.source_url or ""))
    if parsed.scheme == "file":
        # Kept consistent with validate_candidate: without this the two validators
        # disagree, and a locally sourced document passes candidate validation only to
        # fail here, stranding it mid-pipeline instead of being rejected up front.
        if not local_file_sources_allowed():
            errors.append("local_file_source_not_allowed")
    elif parsed.scheme not in {"http", "https"} or not parsed.hostname:
        errors.append("invalid_download_url")
    elif not _is_official_host(source, parsed.hostname):
        errors.append(f"download_redirected_to_non_official_host:{parsed.hostname.lower()}")

    if downloaded.file_size_bytes <= 0 or not downloaded.file_bytes:
        errors.append("empty_download")
        return errors

    extension = _normalize_extension(downloaded.file_ext or downloaded.file_name)
    head = downloaded.file_bytes[:512].lstrip().lower()
    if head.startswith((b"<!doctype html", b"<html")) or b"<title>" in head:
        errors.append("html_response")
    elif extension == ".pdf" and not downloaded.file_bytes.startswith(b"%PDF-"):
        errors.append("invalid_pdf_body")
    elif extension in {".xlsx", ".xlsm", ".zip"} and not downloaded.file_bytes.startswith(b"PK"):
        errors.append("invalid_zip_body")
    elif extension == ".xls" and not downloaded.file_bytes.startswith(b"\xd0\xcf\x11\xe0"):
        errors.append("invalid_xls_body")

    return errors


def validate_parser_smoke(
    downloaded: DownloadedDocument,
    *,
    expected_report_month: date | None = None,
) -> list[str]:
    errors, _detected_report_month = inspect_parser_smoke(
        downloaded,
        expected_report_month=expected_report_month,
    )
    return errors


def inspect_parser_smoke(
    downloaded: DownloadedDocument,
    *,
    expected_report_month: date | None = None,
) -> tuple[list[str], date | None]:
    """Check that a validated body is structurally readable without ingesting it."""
    extension = _normalize_extension(downloaded.file_ext or downloaded.file_name)
    if extension == ".pdf":
        try:
            from pypdf import PdfReader

            reader = PdfReader(BytesIO(downloaded.file_bytes), strict=False)
            if not reader.pages:
                return ["parser_smoke_pdf_no_pages"], None
            if str(downloaded.document_type or "").strip().lower() == "factsheet":
                text = "\n".join(str(page.extract_text() or "") for page in reader.pages)
                detected_report_month = detect_factsheet_content_month(text)
                month_errors = _validate_detected_factsheet_month(
                    detected_report_month,
                    expected_report_month=expected_report_month,
                )
                if month_errors:
                    return month_errors, detected_report_month
                return [], detected_report_month
        except Exception as exc:
            return [f"parser_smoke_pdf_failed:{type(exc).__name__}"], None
    elif extension in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(BytesIO(downloaded.file_bytes), read_only=True, data_only=True)
            if not workbook.sheetnames:
                return ["parser_smoke_excel_no_sheets"], None
            workbook.close()
        except Exception as exc:
            return [f"parser_smoke_excel_failed:{type(exc).__name__}"], None
    elif extension == ".xls":
        # The compound-file signature was checked above. Full XLS parsing remains
        # owned by the existing ingestion parser, which supports its configured engine.
        return [], None
    elif extension == ".zip" and (
        str(downloaded.amc_code or "").strip().upper() == "UTI"
        and str(downloaded.document_type or "").strip().lower() == "portfolio_disclosure"
    ):
        return _inspect_uti_portfolio_zip(
            downloaded.file_bytes,
            expected_report_month=expected_report_month,
        )
    return [], None


def _inspect_uti_portfolio_zip(
    payload: bytes,
    *,
    expected_report_month: date | None,
) -> tuple[list[str], date | None]:
    """Require the UTI consolidated ZIP to contain a real all-scheme portfolio member.

    UTI's monthly archive also carries dividend, derivative, and riskometer files.
    Those are valid spreadsheets but cannot establish holdings readiness. The package
    is accepted only when a member contains both a scheme marker and the official
    portfolio-disclosure date heading.
    """
    try:
        import pandas as pd

        with zipfile.ZipFile(BytesIO(payload)) as archive:
            members = [
                member
                for member in archive.infolist()
                if Path(member.filename).suffix.lower() in {".xls", ".xlsx", ".xlsm"}
                and not member.is_dir()
            ]
            if not members:
                return ["parser_smoke_uti_zip_no_excel_members"], None

            detected_months: set[date] = set()
            for member in members:
                frame = pd.read_excel(
                    BytesIO(archive.read(member.filename)),
                    sheet_name=0,
                    header=None,
                    nrows=250,
                    dtype=str,
                )
                text = "\n".join(
                    " ".join(str(value or "") for value in row)
                    for row in frame.fillna("").values.tolist()
                )
                match = UTI_PORTFOLIO_MONTH_PATTERN.search(text)
                if not match or "scheme:" not in text.lower():
                    continue
                detected_months.add(
                    date(
                        int(match.group("year")),
                        int(match.group("month")),
                        1,
                    )
                )
    except Exception as exc:
        return [f"parser_smoke_uti_zip_failed:{type(exc).__name__}"], None

    if not detected_months:
        return ["parser_smoke_uti_zip_portfolio_member_missing"], None
    if len(detected_months) != 1:
        return ["parser_smoke_uti_zip_multiple_portfolio_months"], None

    detected_month = next(iter(detected_months))
    month_errors = _validate_detected_factsheet_month(
        detected_month,
        expected_report_month=expected_report_month,
    )
    return month_errors, detected_month


def validate_factsheet_content_month(
    text: str,
    *,
    expected_report_month: date | None,
) -> list[str]:
    return _validate_detected_factsheet_month(
        detect_factsheet_content_month(text),
        expected_report_month=expected_report_month,
    )


def detect_factsheet_content_month(text: str) -> date | None:
    from app.mf_ingestion.parsers.factsheet_parser import detect_dominant_factsheet_month

    return detect_dominant_factsheet_month(text)


def _validate_detected_factsheet_month(
    detected: date | None,
    *,
    expected_report_month: date | None,
) -> list[str]:
    if expected_report_month is None:
        return []
    if detected and _month_index(detected) != _month_index(expected_report_month):
        return [
            "factsheet_content_report_month_mismatch:"
            f"{detected.isoformat()}!={expected_report_month.isoformat()}"
        ]
    return []


def content_sha256(downloaded: DownloadedDocument) -> str:
    return sha256(downloaded.file_bytes).hexdigest()


def _is_official_host(source: AMCDocumentSource, hostname: str) -> bool:
    host = hostname.strip().lower().rstrip(".")
    allowed = set(EXTRA_OFFICIAL_HOST_SUFFIXES.get(source.adapter_key.lower(), ()))
    for raw_url in (source.factsheet_page_url, source.portfolio_disclosure_page_url):
        parsed = urlsplit(str(raw_url or ""))
        if parsed.hostname:
            allowed.add(parsed.hostname.lower())

    return any(host == suffix or host.endswith(f".{suffix}") for suffix in allowed)


def _normalize_extension(value: object) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return ""
    if raw.startswith(".") and "/" not in raw:
        return raw
    return Path(raw.split("?", 1)[0]).suffix.lower()


def _month_index(value: date) -> int:
    return value.year * 12 + value.month
